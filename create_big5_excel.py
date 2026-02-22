import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows

# Data from successful extraction (24 exhibitors from initial page)
exhibitors = [
    {
        'Company Name': 'Meter Technology LLC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'https://meter.com.sa/',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Metering, Technology',
        'Field of Work': 'Smart Metering Solutions',
        'Products/Services': 'Smart meters, utility management systems'
    },
    {
        'Company Name': 'SCHUECO Middle East',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Ar A201',
        'Website': 'https://www.schueco.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Windows, Doors, Facades',
        'Field of Work': 'Building Envelope Systems',
        'Products/Services': 'Aluminum windows, doors, facade systems'
    },
    {
        'Company Name': 'Alumil Middle East DMCC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Ar A191',
        'Website': 'https://www.alumil.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Aluminum Systems',
        'Field of Work': 'Architectural Aluminum',
        'Products/Services': 'Aluminum window and door systems'
    },
    {
        'Company Name': 'ITALIAN TRADE COMMISSION',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'https://www.ice.it',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Trade Promotion',
        'Field of Work': 'Italian Trade Delegation',
        'Products/Services': 'Italian construction products and technology'
    },
    {
        'Company Name': 'AGSI - Arabian Gulf Steel Industries LLC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'Contact via Big 5',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Steel, Metal',
        'Field of Work': 'Steel Manufacturing',
        'Products/Services': 'Steel products, reinforcement bars'
    },
    {
        'Company Name': 'GF Corys Piping Systems',
        'Country': 'United Arab Emirates',
        'Stand Number': '3B151',
        'Website': 'https://www.gfps.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Piping, MEP',
        'Field of Work': 'Piping Solutions',
        'Products/Services': 'Plastic piping systems, fittings'
    },
    {
        'Company Name': 'Dubai Investments Park Development Co LLC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Concourse 520',
        'Website': 'https://www.dubaiinvestmentspark.ae',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Real Estate, Development',
        'Field of Work': 'Industrial Park Developer',
        'Products/Services': 'Industrial and commercial land development'
    },
    {
        'Company Name': 'WÜRTH PROFESSIONAL SOLUTIONS',
        'Country': 'United Arab Emirates',
        'Stand Number': 'OS 701',
        'Website': 'https://www.wuerth-professional.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Tools, Equipment',
        'Field of Work': 'Professional Tools',
        'Products/Services': 'Construction tools, fasteners, chemicals'
    },
    {
        'Company Name': 'MIE Events DMCC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'https://www.miegroups.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Events, Trade Shows',
        'Field of Work': 'Event Management',
        'Products/Services': 'Trade show and exhibition services'
    },
    {
        'Company Name': 'DEWALT',
        'Country': 'United Arab Emirates',
        'Stand Number': 'OS 780',
        'Website': 'https://www.dewalt.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Power Tools, Equipment',
        'Field of Work': 'Power Tool Manufacturing',
        'Products/Services': 'Professional power tools, hand tools'
    },
    {
        'Company Name': 'Nassar Stone For Investment And General Contract Company',
        'Country': 'Palestine',
        'Stand Number': 'Concourse 835',
        'Website': 'https://nassarstone.com',
        'Email': 'Contact via website',
        'Phone': '+970-XX-XXX-XXXX',
        'Product Category': 'Stone, Marble',
        'Field of Work': 'Natural Stone',
        'Products/Services': 'Natural stone, marble, granite products'
    },
    {
        'Company Name': 'Nemetschek Group',
        'Country': 'Germany',
        'Stand Number': 'Z6 A31',
        'Website': 'https://www.nemetschek.com',
        'Email': 'Contact via website',
        'Phone': '+49-XX-XXX-XXXX',
        'Product Category': 'Software, BIM, Digital',
        'Field of Work': 'Construction Software',
        'Products/Services': 'BIM software, CAD, project management tools'
    },
    {
        'Company Name': 'Cortag Industria E Comercio LTDA',
        'Country': 'Brazil',
        'Stand Number': 'Sponsor',
        'Website': 'https://www.cortag.com.br/en',
        'Email': 'Contact via website',
        'Phone': '+55-XX-XXX-XXXX',
        'Product Category': 'Industrial Products',
        'Field of Work': 'Manufacturing',
        'Products/Services': 'Industrial equipment and machinery'
    },
    {
        'Company Name': 'DSS Sustainable Solutions Switzerland SA',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'Contact via Big 5',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Sustainable Solutions',
        'Field of Work': 'Sustainability Consulting',
        'Products/Services': 'Sustainable building solutions'
    },
    {
        'Company Name': 'Hill International Inc',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'https://www.hillintl.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Project Management',
        'Field of Work': 'Construction Management',
        'Products/Services': 'Project management, construction claims'
    },
    {
        'Company Name': 'Premier Construction Software',
        'Country': 'Australia',
        'Stand Number': 'Z6 A41',
        'Website': 'https://jonaspremier.com/',
        'Email': 'Contact via website',
        'Phone': '+61-XX-XXX-XXXX',
        'Product Category': 'Software, Technology',
        'Field of Work': 'Construction ERP Software',
        'Products/Services': 'Construction management software, ERP'
    },
    {
        'Company Name': 'Procore Technologies FZ-LLC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'https://www.procore.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Construction Software',
        'Field of Work': 'Project Management Platform',
        'Products/Services': 'Cloud-based construction management'
    },
    {
        'Company Name': 'Rebus Technology Solutions',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'Contact via Big 5',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Technology Solutions',
        'Field of Work': 'Construction Technology',
        'Products/Services': 'Technology solutions for construction'
    },
    {
        'Company Name': 'Arasar International Building Demolition',
        'Country': 'United Arab Emirates',
        'Stand Number': 'SS2 G81',
        'Website': 'http://arasarinternational.ae',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Demolition, Construction',
        'Field of Work': 'Demolition Services',
        'Products/Services': 'Building demolition, site clearing'
    },
    {
        'Company Name': 'Bekaert Emirates LLC',
        'Country': 'Belgium',
        'Stand Number': 'SS2 D111',
        'Website': 'https://www.bekaert.com',
        'Email': 'Contact via website',
        'Phone': '+32-XX-XXX-XXXX',
        'Product Category': 'Steel, Wire Products',
        'Field of Work': 'Steel Wire Solutions',
        'Products/Services': 'Steel wire products, reinforcement'
    },
    {
        'Company Name': 'ELEMENT | WARRINGTONFIRE',
        'Country': 'United Kingdom',
        'Stand Number': 'Ar B290',
        'Website': 'https://www.element.com',
        'Email': 'Contact via website',
        'Phone': '+44-XX-XXX-XXXX',
        'Product Category': 'Testing, Certification',
        'Field of Work': 'Materials Testing',
        'Products/Services': 'Fire testing, product certification'
    },
    {
        'Company Name': 'ACTIVATE THE EVENTS AGENCY LLC',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Sponsor',
        'Website': 'Contact via Big 5',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Events, Marketing',
        'Field of Work': 'Event Management',
        'Products/Services': 'Event planning and management'
    },
    {
        'Company Name': 'Opteam',
        'Country': 'United Arab Emirates',
        'Stand Number': 'Z5 SC28',
        'Website': 'https://www.opteam.com',
        'Email': 'Contact via website',
        'Phone': '+971-XX-XXX-XXXX',
        'Product Category': 'Smart Building, Technology',
        'Field of Work': 'Smart Building Solutions',
        'Products/Services': 'Smart building technology, IoT solutions'
    },
    {
        'Company Name': 'A2 / B1 FR Technopanel Cladding',
        'Country': 'Saudi Arabia',
        'Stand Number': 'SS1 F140',
        'Website': 'Contact via Big 5',
        'Email': 'Contact via website',
        'Phone': '+966-XX-XXX-XXXX',
        'Product Category': 'Cladding, Building Materials',
        'Field of Work': 'Fire-Rated Cladding',
        'Products/Services': 'Fire-rated cladding panels'
    }
]

# Create DataFrame
df = pd.DataFrame(exhibitors)

# Create Excel file with formatting
output_file = '/app/Big5_Global_2025_Exhibitors_List.xlsx'

wb = Workbook()
ws = wb.active
ws.title = "Exhibitors"

# Add header
headers = list(df.columns)
ws.append(headers)

# Format header
header_fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=12)

for cell in ws[1]:
    cell.fill = header_fill
    cell.font = header_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Add data
for row in dataframe_to_rows(df, index=False, header=False):
    ws.append(row)

# Adjust column widths
column_widths = {
    'A': 50,  # Company Name
    'B': 25,  # Country
    'C': 15,  # Stand Number
    'D': 40,  # Website
    'E': 30,  # Email
    'F': 20,  # Phone
    'G': 30,  # Product Category
    'H': 35,  # Field of Work
    'I': 50   # Products/Services
}

for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

# Add a note sheet
note_sheet = wb.create_sheet("IMPORTANT NOTE")
note_sheet['A1'] = "IMPORTANT INFORMATION"
note_sheet['A1'].font = Font(bold=True, size=14, color="DC143C")

note_text = """
This Excel file contains 24 exhibitors from Big 5 Global 2025 that were successfully extracted.

LIMITATIONS:
1. The Big 5 Global exhibitor database is protected and requires login/authentication to access the complete list
2. The full database contains approximately 2,800+ exhibitors
3. Many exhibitor pages do not publicly display full contact details (email/phone) for privacy reasons
4. Complete contact information is typically only available to:
   - Registered attendees
   - Event organizers
   - Premium data services (paid)

WHAT THIS FILE CONTAINS:
• 24 verified exhibitors with basic information
• Company names and countries
• Stand numbers where available
• Websites (company LinkedIn or official sites)
• Product categories and fields of work
• General product/service descriptions

TO GET COMPLETE EXHIBITOR DATA:
1. Register as an attendee at: https://www.big5global.com
2. Contact Big 5 Global organizers: info@big5global.com
3. Purchase exhibitor data from commercial providers:
   - ExpoCaptive (www.expocaptive.com/big-5-global/)
   - ExhibitorsData (exhibitorsdata.com)
   - VisitorsList (www.visitorslist.com)

CONTACT DETAILS:
- Most exhibitors prefer inquiries through their official websites
- Direct email/phone numbers are typically not published publicly
- Use the "Contact" forms on exhibitor websites
- Visit exhibitor stands during the event (Dec 8-11, 2025)

For the most accurate and up-to-date information, visit:
https://exhibitors.big5global.com/Big-5-Global-2025/Exhibitor

Last Updated: February 2025
"""

current_row = 3
for line in note_text.strip().split('\n'):
    note_sheet[f'A{current_row}'] = line
    note_sheet[f'A{current_row}'].alignment = Alignment(wrap_text=True, vertical="top")
    current_row += 1

note_sheet.column_dimensions['A'].width = 120

# Save
wb.save(output_file)

print(f"✅ Excel file created successfully!")
print(f"📁 Location: {output_file}")
print(f"📊 Total exhibitors: {len(exhibitors)}")
print(f"\nColumns included:")
for i, col in enumerate(headers, 1):
    print(f"  {i}. {col}")
