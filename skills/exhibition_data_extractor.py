"""
EXHIBITION DATA EXTRACTOR - Universal Tool
==========================================
A reusable skill to login and extract exhibitor data from any exhibition website
and create comprehensive Excel files with all details.

Author: AI Assistant
Version: 1.0
"""

import asyncio
import json
import re
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExhibitionDataExtractor:
    """
    Universal Exhibition Data Extractor
    Handles login, browsing, and data extraction from exhibition websites
    """
    
    def __init__(self):
        self.exhibitors = []
        self.config = {
            'user_agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36',
            'timeout': 60000,
            'wait_after_action': 2000,
            'scroll_times': 30,
            'max_exhibitors': 5000
        }
    
    def clean_text(self, text):
        """Clean and normalize text"""
        if not text:
            return ""
        return ' '.join(text.split()).strip()
    
    def extract_email(self, text):
        """Extract email from text"""
        emails = re.findall(r'\b[\w\.-]+@[\w\.-]+\.\w{2,}\b', text)
        # Filter out common non-contact emails
        blacklist = ['example.com', 'test.com', 'placeholder.com', 'domain.com']
        valid_emails = [e for e in emails if not any(b in e.lower() for b in blacklist)]
        return valid_emails[0] if valid_emails else ""
    
    def extract_phone(self, text):
        """Extract phone number from text"""
        patterns = [
            r'\+\d{1,4}[\s\-]?\(?\d{1,4}\)?[\s\-]?\d{1,4}[\s\-]?\d{4,9}',
            r'\(\d{3}\)\s?\d{3}[\s\-]?\d{4}',
            r'\d{3}[\s\-]\d{3}[\s\-]\d{4}'
        ]
        for pattern in patterns:
            phones = re.findall(pattern, text)
            if phones:
                # Filter out years
                valid = [p for p in phones if not any(year in p for year in ['2024', '2025', '2026'])]
                if valid:
                    return valid[0]
        return ""
    
    def extract_website(self, links, exclude_domains=None):
        """Extract valid website from list of links"""
        if exclude_domains is None:
            exclude_domains = ['facebook.com', 'twitter.com', 'instagram.com', 'youtube.com']
        
        for link in links:
            if link and link.startswith('http'):
                if not any(domain in link for domain in exclude_domains):
                    return link
        return ""
    
    async def login_to_website(self, page, url, username, password, selectors=None):
        """
        Generic login function for exhibition websites
        
        Args:
            page: Playwright page object
            url: Login URL
            username: Username/email for login
            password: Password for login
            selectors: Dict with login form selectors (optional, will auto-detect if None)
        """
        print(f"🔐 Attempting login to: {url}")
        
        try:
            await page.goto(url, wait_until="networkidle", timeout=self.config['timeout'])
            await page.wait_for_timeout(3000)
            
            if selectors:
                # Use provided selectors
                await page.fill(selectors['username'], username)
                await page.fill(selectors['password'], password)
                await page.click(selectors['submit'])
            else:
                # Auto-detect login form
                print("  Auto-detecting login form...")
                
                # Try common username/email field patterns
                username_selectors = [
                    'input[type="email"]',
                    'input[name="email"]',
                    'input[name="username"]',
                    'input[id*="email"]',
                    'input[id*="username"]',
                    'input[placeholder*="email"]',
                    'input[placeholder*="Email"]'
                ]
                
                username_filled = False
                for selector in username_selectors:
                    try:
                        if await page.query_selector(selector):
                            await page.fill(selector, username)
                            username_filled = True
                            print(f"  ✓ Filled username field: {selector}")
                            break
                    except:
                        continue
                
                if not username_filled:
                    print("  ✗ Could not find username field")
                    return False
                
                # Try common password field patterns
                password_selectors = [
                    'input[type="password"]',
                    'input[name="password"]',
                    'input[id*="password"]'
                ]
                
                password_filled = False
                for selector in password_selectors:
                    try:
                        if await page.query_selector(selector):
                            await page.fill(selector, password)
                            password_filled = True
                            print(f"  ✓ Filled password field: {selector}")
                            break
                    except:
                        continue
                
                if not password_filled:
                    print("  ✗ Could not find password field")
                    return False
                
                # Try common submit button patterns
                submit_selectors = [
                    'button[type="submit"]',
                    'input[type="submit"]',
                    'button:has-text("Login")',
                    'button:has-text("Sign in")',
                    'button:has-text("Log in")',
                    'a:has-text("Login")'
                ]
                
                submit_clicked = False
                for selector in submit_selectors:
                    try:
                        if await page.query_selector(selector):
                            await page.click(selector)
                            submit_clicked = True
                            print(f"  ✓ Clicked submit button: {selector}")
                            break
                    except:
                        continue
                
                if not submit_clicked:
                    # Try pressing Enter as fallback
                    await page.keyboard.press('Enter')
                    print("  ✓ Pressed Enter to submit")
            
            # Wait for navigation after login
            await page.wait_for_timeout(5000)
            
            # Check if login was successful
            current_url = page.url
            page_content = await page.content()
            
            # Login success indicators
            success_indicators = ['dashboard', 'logout', 'sign out', 'my account', 'profile']
            failure_indicators = ['login failed', 'invalid credentials', 'incorrect password', 'error']
            
            page_text_lower = page_content.lower()
            
            if any(indicator in page_text_lower for indicator in success_indicators):
                print("  ✅ Login successful!")
                return True
            elif any(indicator in page_text_lower for indicator in failure_indicators):
                print("  ❌ Login failed - invalid credentials")
                return False
            else:
                print("  ⚠️  Login status unclear, proceeding with caution...")
                return True
                
        except Exception as e:
            print(f"  ❌ Login error: {str(e)}")
            return False
    
    async def extract_exhibitors_from_page(self, page, extraction_config=None):
        """
        Extract exhibitor data from current page
        
        Args:
            page: Playwright page object
            extraction_config: Custom extraction configuration (optional)
        """
        print("📊 Extracting exhibitor data from current page...")
        
        # Scroll to load all content
        print(f"  Scrolling to load all exhibitors ({self.config['scroll_times']} times)...")
        for i in range(self.config['scroll_times']):
            await page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
            await page.wait_for_timeout(1500)
            if (i + 1) % 10 == 0:
                print(f"    Progress: {i+1}/{self.config['scroll_times']}")
        
        print("  Extracting data...")
        
        # Extract exhibitor data using JavaScript
        exhibitors_data = await page.evaluate('''() => {
            const exhibitors = [];
            const seen = new Set();
            
            // Common selectors for exhibitor containers
            const containerSelectors = [
                '.exhibitor-card',
                '.exhibitor-item',
                '.exb-list-item',
                '[class*="exhibitor"]',
                '[class*="company"]',
                '.company-card',
                '.vendor-card',
                'article',
                '.card'
            ];
            
            let containers = [];
            for (const selector of containerSelectors) {
                const elements = document.querySelectorAll(selector);
                if (elements.length > 0) {
                    containers = Array.from(elements);
                    break;
                }
            }
            
            containers.forEach(container => {
                const exhibitor = {
                    name: '',
                    country: '',
                    standNo: '',
                    website: '',
                    email: '',
                    phone: '',
                    address: '',
                    category: '',
                    products: '',
                    description: '',
                    logo: '',
                    contactPerson: ''
                };
                
                // Extract company name
                const nameSelectors = ['h1', 'h2', 'h3', 'h4', 'h5', '.company-name', '.exhibitor-name', '[class*="name"]', '[class*="title"]'];
                for (const sel of nameSelectors) {
                    const elem = container.querySelector(sel);
                    if (elem && elem.textContent.trim().length > 2) {
                        exhibitor.name = elem.textContent.trim();
                        break;
                    }
                }
                
                // Get all text content
                const text = container.textContent;
                
                // Extract country
                const countries = [
                    'United Arab Emirates', 'UAE', 'Saudi Arabia', 'China', 'India', 'Germany', 'Italy',
                    'United Kingdom', 'UK', 'France', 'Spain', 'Turkey', 'Egypt', 'Brazil', 'Canada',
                    'Australia', 'Japan', 'South Korea', 'Netherlands', 'Belgium', 'Switzerland',
                    'Austria', 'Poland', 'Russia', 'Pakistan', 'Thailand', 'Vietnam', 'Indonesia',
                    'Malaysia', 'Singapore', 'Hong Kong', 'Taiwan', 'Lebanon', 'Jordan', 'Kuwait',
                    'Qatar', 'Bahrain', 'Oman', 'Morocco', 'Greece', 'Portugal', 'Sweden', 'Norway',
                    'Denmark', 'Finland', 'Ireland', 'New Zealand', 'South Africa'
                ];
                
                for (const country of countries) {
                    if (text.includes(country)) {
                        exhibitor.country = country;
                        break;
                    }
                }
                
                // Extract stand number
                const standPatterns = [
                    /Stand\s+(?:No[:\.\-\s]+)?([A-Z]{1,3}\s*[A-Z]?\d{1,4})/i,
                    /Hall\s+(\d+[A-Z]?)/i,
                    /Booth\s+([A-Z]?\d+)/i,
                    /Concourse\s+(\d+)/i
                ];
                
                for (const pattern of standPatterns) {
                    const match = text.match(pattern);
                    if (match) {
                        exhibitor.standNo = match[1].trim();
                        break;
                    }
                }
                
                // Extract website
                const links = container.querySelectorAll('a[href^="http"]');
                const excludeDomains = ['facebook.com', 'twitter.com', 'instagram.com', 'youtube.com', 'linkedin.com'];
                
                for (const link of links) {
                    const href = link.href;
                    if (!excludeDomains.some(d => href.includes(d))) {
                        exhibitor.website = href;
                        break;
                    }
                }
                
                // Extract email
                const emailMatch = text.match(/[\w\.-]+@[\w\.-]+\.\w{2,}/);
                if (emailMatch) {
                    exhibitor.email = emailMatch[0];
                }
                
                // Extract phone
                const phoneMatch = text.match(/\+?\d{1,4}[\s\-]?\(?\d{1,4}\)?[\s\-]?\d{1,4}[\s\-]?\d{4,}/);
                if (phoneMatch) {
                    exhibitor.phone = phoneMatch[0];
                }
                
                // Extract logo
                const img = container.querySelector('img');
                if (img && img.src && !img.src.includes('placeholder')) {
                    exhibitor.logo = img.src;
                }
                
                // Extract description
                const descSelectors = ['.description', '.about', '[class*="desc"]', 'p'];
                for (const sel of descSelectors) {
                    const elem = container.querySelector(sel);
                    if (elem && elem.textContent.trim().length > 50) {
                        exhibitor.description = elem.textContent.trim().substring(0, 300);
                        break;
                    }
                }
                
                // Add to list if has name
                if (exhibitor.name && exhibitor.name.length > 2) {
                    const key = exhibitor.name + exhibitor.country;
                    if (!seen.has(key)) {
                        seen.add(key);
                        exhibitors.push(exhibitor);
                    }
                }
            });
            
            return exhibitors;
        }''')
        
        print(f"  ✅ Extracted {len(exhibitors_data)} exhibitors from current page")
        return exhibitors_data
    
    async def navigate_pagination(self, page, max_pages=100):
        """
        Navigate through pagination to get all exhibitors
        
        Args:
            page: Playwright page object
            max_pages: Maximum pages to scrape (default 100)
        """
        all_exhibitors = []
        current_page = 1
        
        print(f"🔄 Navigating through pages (max {max_pages} pages)...")
        
        while current_page <= max_pages:
            print(f"\n📄 Processing page {current_page}/{max_pages}")
            
            # Extract exhibitors from current page
            exhibitors = await self.extract_exhibitors_from_page(page)
            all_exhibitors.extend(exhibitors)
            
            # Try to find and click next page button
            next_selectors = [
                'a:has-text("Next")',
                'button:has-text("Next")',
                'a:has-text("›")',
                'a:has-text("→")',
                '.pagination .next',
                '[aria-label="Next"]',
                'a.next-page'
            ]
            
            next_clicked = False
            for selector in next_selectors:
                try:
                    next_button = await page.query_selector(selector)
                    if next_button:
                        await next_button.click()
                        await page.wait_for_timeout(3000)
                        next_clicked = True
                        print(f"  ✓ Clicked next page button")
                        break
                except:
                    continue
            
            if not next_clicked:
                print(f"  ℹ️  No more pages found. Stopping at page {current_page}")
                break
            
            current_page += 1
            
            # Safety check - stop if we have enough
            if len(all_exhibitors) >= self.config['max_exhibitors']:
                print(f"  ℹ️  Reached max exhibitors limit ({self.config['max_exhibitors']})")
                break
        
        return all_exhibitors
    
    def create_excel_file(self, exhibitors, output_file, event_name="Exhibition"):
        """
        Create formatted Excel file with exhibitor data
        
        Args:
            exhibitors: List of exhibitor dictionaries
            output_file: Output file path
            event_name: Name of the event
        """
        print(f"\n📊 Creating Excel file: {output_file}")
        
        # Create DataFrame
        df = pd.DataFrame(exhibitors)
        
        # Reorder columns
        column_order = [
            'name', 'country', 'standNo', 'website', 'email', 'phone',
            'category', 'products', 'description', 'address', 'contactPerson', 'logo'
        ]
        
        # Only include columns that exist
        available_columns = [col for col in column_order if col in df.columns]
        df = df[available_columns]
        
        # Rename columns to be more readable
        column_names = {
            'name': 'Company Name',
            'country': 'Country',
            'standNo': 'Stand Number',
            'website': 'Website',
            'email': 'Email',
            'phone': 'Phone',
            'category': 'Product Category',
            'products': 'Products/Services',
            'description': 'Description',
            'address': 'Address',
            'contactPerson': 'Contact Person',
            'logo': 'Logo URL'
        }
        
        df = df.rename(columns=column_names)
        
        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Exhibitors"
        
        # Add title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f"{event_name} - Exhibitor List"
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="DC143C", end_color="DC143C", fill_type="solid")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add extraction info
        ws.merge_cells('A2:F2')
        info_cell = ws['A2']
        info_cell.value = f"Extracted: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} | Total: {len(exhibitors)} companies"
        info_cell.font = Font(italic=True, size=10)
        info_cell.alignment = Alignment(horizontal="center")
        
        # Add headers (row 3)
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_idx)
            cell.value = header
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # Add data
        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.alignment = Alignment(vertical="top", wrap_text=True)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
        
        # Auto-adjust column widths
        column_widths = {
            'Company Name': 40,
            'Country': 20,
            'Stand Number': 15,
            'Website': 40,
            'Email': 30,
            'Phone': 18,
            'Product Category': 30,
            'Products/Services': 45,
            'Description': 50,
            'Address': 35,
            'Contact Person': 25,
            'Logo URL': 40
        }
        
        for col_idx, column_name in enumerate(df.columns, 1):
            if column_name in column_widths:
                ws.column_dimensions[chr(64 + col_idx)].width = column_widths[column_name]
        
        # Freeze header rows
        ws.freeze_panes = 'A4'
        
        # Save workbook
        wb.save(output_file)
        
        print(f"✅ Excel file created successfully!")
        print(f"   📁 Location: {output_file}")
        print(f"   📊 Total exhibitors: {len(exhibitors)}")
        print(f"   📋 Columns: {len(df.columns)}")
        
        # Print statistics
        print(f"\n📈 Statistics:")
        print(f"   • With Email: {df['Email'].notna().sum()}")
        print(f"   • With Phone: {df['Phone'].notna().sum()}")
        print(f"   • With Website: {df['Website'].notna().sum()}")
        if 'Country' in df.columns:
            print(f"   • Countries: {df['Country'].nunique()}")
        
        return output_file


# Example usage function
async def extract_from_exhibition_site(
    url,
    output_file,
    event_name,
    login_required=False,
    username=None,
    password=None,
    login_url=None
):
    """
    Main function to extract exhibition data
    
    Args:
        url: Exhibition exhibitor list URL
        output_file: Output Excel file path
        event_name: Name of the event
        login_required: Whether login is required
        username: Username for login (if required)
        password: Password for login (if required)
        login_url: Login page URL (if different from main URL)
    """
    print("="*70)
    print(f"EXHIBITION DATA EXTRACTOR")
    print(f"Event: {event_name}")
    print(f"URL: {url}")
    print("="*70)
    
    extractor = ExhibitionDataExtractor()
    
    # This would be called by the Playwright screenshot tool
    # Example of how to use it is in the next file
    
    return extractor


if __name__ == "__main__":
    print("Exhibition Data Extractor - Universal Tool")
    print("This module should be imported and used with Playwright")
    print("See usage_example.py for implementation details")
